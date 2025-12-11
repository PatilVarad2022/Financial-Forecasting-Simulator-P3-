"""
Excel Formula Audit Script
Verifies that critical sheets contain formulas, not just values
"""
from openpyxl import load_workbook
import os

def audit_excel_formulas(filepath):
    """Check if Excel sheets contain formulas"""
    if not os.path.exists(filepath):
        print(f"❌ File not found: {filepath}")
        return False
    
    print(f"📊 Auditing: {filepath}\n")
    
    wb = load_workbook(filepath, data_only=False)
    sheets_to_check = ['Engine', 'Working_Capital', 'Depreciation_Sched', 'Sensitivity', 'Checks']
    
    all_passed = True
    
    for sheet_name in sheets_to_check:
        if sheet_name not in wb.sheetnames:
            print(f"⚠️  Sheet '{sheet_name}' not found")
            continue
            
        ws = wb[sheet_name]
        found_formulas = False
        formula_samples = []
        
        for row in ws.iter_rows(min_row=1, max_row=50, min_col=1, max_col=20):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    found_formulas = True
                    if len(formula_samples) < 3:
                        formula_samples.append(f"{cell.coordinate}: {cell.value[:60]}...")
        
        status = "✅" if found_formulas else "❌"
        print(f"{status} {sheet_name}: formulas found? {found_formulas}")
        
        if formula_samples:
            print(f"   Sample formulas:")
            for sample in formula_samples:
                print(f"     - {sample}")
        
        if not found_formulas:
            all_passed = False
        
        print()
    
    wb.close()
    
    if all_passed:
        print("✅ All critical sheets contain formulas!")
    else:
        print("❌ Some sheets are value-only. Model needs rebuilding.")
    
    return all_passed

if __name__ == '__main__':
    filepath = 'outputs/FPnA_Model_with_formulas.xlsx'
    audit_excel_formulas(filepath)
